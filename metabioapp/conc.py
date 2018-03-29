import json
from numpy.polynomial import polynomial as poly
import openpyxl
from numpy import array, mean, std


def is_metabolite(name):
    if name != 'id' and name != 'Concentration' and name != 'Group' and name != 'std_replicate' and name != 'NF' and not name.startswith('IS'):
        return True
    return False

def normalise_self(is_list):
    min_val = min(is_list)
    if min_val == 0:
        min_val = 1
    for i in range(0, len(is_list)):
        is_list[i] /= min_val

def cal_equation(y, equation, origin):
    if origin == 'default':
        if len(equation) == 2:
            return (y - equation[0]) / equation[1]
        if len(equation) == 3:
            delta = equation[1]**2 - 4*equation[2]*(equation[0] - y)
            if delta < 0:
                return 0
            x1 = (-equation[1] + delta ** (1.0/2)) / 2*equation[2]
            x2 = (-equation[1] - delta ** (1.0/2)) / 2*equation[2]
            if x1 > 0:
                return x1
            elif x2 > 0:
                return x2
            else:
                return 0
    if origin == 'forced':
        if len(equation) == 2:
            return y / equation[1]
        if len(equation) == 3:
            delta = equation[1]**2 + 4 * equation[3] * y
            if delta < 0:
                return 0
            x1 = (-equation[1] + delta ** (1.0 / 2)) / 2 * equation[2]
            x2 = (-equation[1] - delta ** (1.0 / 2)) / 2 * equation[2]
            if x1 > 0:
                return x1
            elif x2 > 0:
                return x2
            else:
                return 0
    return 0

def cal_predict_points(points, equation, origin):
    res = []
    '''
    if origin == 'default':
        if len(equation) == 3:
            for each in points:
                res.append(equation[0] + each*equation[1] + each*each*equation[2])
            return res
        elif len(equation) == 2:
            for each in points:
                res.append(equation[0] + each*equation[1])
    if origin == 'forced':
        if len(equation) == 3:
            for each in points:
                res.append(each*equation[1] + each*each*equation[2])
            return res
        elif len(equation) == 2:
            for each in points:
                res.append(each*equation[1])
    '''
    for y in points:
        res.append(cal_equation(y, equation, origin))
    return res

def is_normalisation(input_dic, is_method_dic):
    is_dic = {}
    # store all IS column in a dic structure:  {'IS1' : [1,2,3,4,6]}
    for each_col in input_dic:
        if each_col.startswith('IS'):
            is_dic[each_col] = input_dic[each_col]
    # normalise itself
    for each_list in is_dic.values():
        normalise_self(each_list)
    #divide all the metabolite values by the corresponding IS values
    for each_col in input_dic:
        if is_metabolite(each_col):
            if is_method_dic[each_col] == 'None':
                continue
            #metabolite_values = input_dic[each_col]
            is_values = is_dic[is_method_dic[each_col]]
            for i in range(0, len(input_dic[each_col])):
                if is_values[i] == 0:
                    is_values[i] = 1
                input_dic[each_col][i] /= is_values[i]
    return input_dic

def reagent_sub(input_dic):
    for each_col in input_dic:
        total = 0
        count = 0
        if is_metabolite(each_col):
            for i in range(len(input_dic[each_col])):
                if input_dic['Group'][i] == 'R':
                    total += input_dic[each_col][i]
                    count += 1
            ave = total / count

            for i in range(len(input_dic[each_col])):
                input_dic[each_col][i] -= ave
    return input_dic

def cal_regression(input_dic, regression_option):
    for each_col in input_dic:
        if is_metabolite(each_col):
            method = regression_option[each_col][0]
            origin = regression_option[each_col][1]
            weight = regression_option[each_col][2]
            points = []
            weights = []
            dataX = []
            dataY = []
            equation = []
            for i in range(0, len(input_dic[each_col])):
                weights = []
                if input_dic['Group'][i] == 'S':
                    tmp = []
                    tmp.append(input_dic['Concentration'][i])
                    tmp.append(input_dic[each_col][i])
                    #points.append([input_dic['Concentration'][i], input_dic[each_col][i]])
                    points.append(tmp)
            for each_point in points:
                if weight is None or weight == 'none':
                    weights.append(1)
                if weight == '1/x':
                    if each_point[0] == 0:
                        weights.append(0)
                    else:
                        weights.append(1 / each_point[0])
                if weight == '1/x^2':
                    if each_point[0] == 0:
                        weights.append(0)
                    else:
                        weights.append(1 / (each_point[0] ** 2))
                if weight == '1/y':
                    if each_point[1] == 0:
                        weights.append(0)
                    else:
                        weights.append(1 / each_point[1])
                if weight == '1/y^2':
                    if each_point[1] == 0:
                        weights.append(0)
                    else:
                        weights.append(1 / (each_point[1] ** 2))

                dataX.append(each_point[0])
                dataY.append(each_point[1])
            if method == 'quad':
                equation, stats = poly.polyfit(x=dataX, y=dataY, deg=2, w=weights, full=True)
            elif method == 'linear':
                equation, stats = poly.polyfit(x=dataX, y=dataY, deg=1, w=weights, full=True)
            res = cal_predict_points(input_dic[each_col], equation, origin)
            for i in range(0, len(res)):
                input_dic[each_col][i] = res[i]
    return input_dic

def write_data(input_dic, workbook, sheet_name, text, dilution_factor):
    #workbook = openpyxl.load_workbook(workbook_path)
    cur_sheet = workbook.create_sheet(title=sheet_name)
    cur_sheet.append([text])
    head = []
    for col_name in input_dic:
        head.append(col_name)
    cur_sheet.append(head)
    for i in range(0, len(input_dic['id'])):
        tmp = []
        for each_col in input_dic:
            if each_col == 'id' or each_col == 'Group':
                tmp.append(input_dic[each_col][i])
            else:
                tmp.append(input_dic[each_col][i] * dilution_factor)
        cur_sheet.append(tmp)
    #workbook.save(workbook_path)

def stats(input_dic, workbook, sheet_name, text):
    group_index = {}
    index = 0
    for group_name in input_dic['Group']:
        if group_name in group_index:
            group_index[group_name].append(index)
        else:
            group_index[group_name] = [index]
        index += 1

    #workbook = openpyxl.load_workbook(workbook_path)
    cur_sheet = workbook.create_sheet(title=sheet_name)
    cur_sheet.append([text])
    cur_sheet.append([])
    for each_group in group_index:
        cur_sheet.append(['In Group ' + each_group])
        for each_col in input_dic:
            if is_metabolite(each_col):
                tmp = []
                indexes = group_index[each_group]
                for index in indexes:
                    tmp.append(input_dic[each_col][index])
                tmp_array = array(tmp)
                if mean(tmp_array) == 0:
                    cur_sheet.append([each_col, 'NA'])
                else:
                    cur_sheet.append([each_col, str(std(tmp_array) / mean(tmp_array) * 100)])

        cur_sheet.append([])
    #workbook.save(workbook_path)

def normalising_factor(input_dic):
    nf = input_dic['NF']
    for each_col in input_dic:
        if is_metabolite(each_col):
            for i in range(0, len(input_dic[each_col])):
                input_dic[each_col][i] /= nf[i]
    return input_dic


def generate_output(inputdict, path, options):
    input_dic = inputdict
    is_method_dic = json.loads(options['IS_method'])
    need_is = options['need_IS']
    need_rea = options['need_rea']
    regression_option = json.loads(options['regression_option'])
    dilution_factor = float(options['dilution_factor'])
    nf = options['nf']

    workbook = openpyxl.Workbook()

    write_data(input_dic, workbook, 'raw', 'Raw Data', 1)  #write raw data


    if need_is == 'yes':
        input_dic = is_normalisation(input_dic, is_method_dic)
        write_data(input_dic, workbook, 'IS', 'IS Normalised Data', 1)
        stats(input_dic, workbook, 'CV_IS', 'CVs After Internal Standard Normalisation')
    else:
        stats(input_dic, workbook, 'CV', 'CVs')

    if need_rea == 'yes':
        input_dic = reagent_sub(input_dic)
        if need_is == 'yes':
            write_data(input_dic, workbook, 'IS_REA', 'IS Normalised, Reagent Blank Subtracted Data', 1)
            stats(input_dic, workbook, 'CV_IS_REA', 'CVs after IS Normalised and Reagent Blank Subtraction')
        else:
            write_data(input_dic, workbook, 'REA', 'Reagent Blank Subtracted Data', 1)
            stats(input_dic, workbook, 'CV_REA', 'CVs after Reagent Blank Subtraction')
    input_dic = cal_regression(input_dic, regression_option)
    if nf == 'yes':
        input_dic = normalising_factor(input_dic)
    write_data(input_dic, workbook, 'Con', 'Concentration Data', 1)
    write_data(input_dic, workbook, 'Con_Dil', 'Concentration Data with Dilution Factor', dilution_factor)
    stats(input_dic, workbook, 'CV_Con', 'CVs of Concentrations')

    active_before = workbook.active
    sheet_names = []
    for sheet in workbook.worksheets:
        sheet_names.append(sheet.title)
        if sheet.title == 'Sheet':
            workbook.remove(sheet)
    active_after = workbook.active

    workbook.save(path)





